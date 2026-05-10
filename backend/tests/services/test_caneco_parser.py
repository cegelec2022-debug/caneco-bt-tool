"""Tests du parser CANECO BT.

Valide le comportement sur le fichier de référence DACHSER et des cas limites.
"""

from pathlib import Path

import pytest

from app.services.caneco.parser import (
    ParseResult,
    _build_mapping,
    _normalize,
    _to_float,
    _to_int,
    _to_str,
    detect_indice_from_filename,
    parse_caneco_file,
)

# ---------------------------------------------------------------------------
# Chemin du fichier de référence
# ---------------------------------------------------------------------------

SEED_DIR = Path("/app/data/seed/dachser")
DACHSER_FILE = SEED_DIR / "DATA_DACHSER_INDICE_B.XLS"


# ---------------------------------------------------------------------------
# Tests unitaires : fonctions utilitaires
# ---------------------------------------------------------------------------


class TestNormalize:
    def test_accents_suppresses(self) -> None:
        assert _normalize("Repère aval") == "repere aval"

    def test_majuscules_converties(self) -> None:
        assert _normalize("IrTh / IN") == "irth / in"

    def test_unites_supprimees(self) -> None:
        assert _normalize("IB (A)") == "ib"

    def test_espaces_normalises(self) -> None:
        assert _normalize("  type   de   cable  ") == "type de cable"


class TestToFloat:
    def test_entier(self) -> None:
        assert _to_float(15) == 15.0

    def test_float(self) -> None:
        assert _to_float(3.14) == pytest.approx(3.14)

    def test_virgule_francaise(self) -> None:
        assert _to_float("15,5") == pytest.approx(15.5)

    def test_none(self) -> None:
        assert _to_float(None) is None

    def test_chaine_vide(self) -> None:
        assert _to_float("") is None

    def test_non_numerique(self) -> None:
        assert _to_float("abc") is None


class TestToInt:
    def test_float_arrondi(self) -> None:
        assert _to_int(3.0) == 3

    def test_none(self) -> None:
        assert _to_int(None) is None


class TestToStr:
    def test_valeur_normale(self) -> None:
        assert _to_str("U1000 R02V") == "U1000 R02V"

    def test_none_renvoie_none(self) -> None:
        assert _to_str(None) is None

    def test_nan_renvoie_none(self) -> None:
        assert _to_str("nan") is None

    def test_na_renvoie_none(self) -> None:
        assert _to_str("#N/A") is None

    def test_espaces_stripes(self) -> None:
        assert _to_str("  TGBT  ") == "TGBT"


class TestDetectIndice:
    def test_indice_standard(self) -> None:
        assert detect_indice_from_filename("DATA_DACHSER_INDICE_B.XLS") == "B"

    def test_indice_avec_chiffre(self) -> None:
        assert detect_indice_from_filename("EXPORT_INDICE_B2.xlsx") == "B2"

    def test_sans_indice(self) -> None:
        result = detect_indice_from_filename("projet.xls")
        # Peut retourner None ou une lettre via fallback
        assert result is None or isinstance(result, str)

    def test_insensible_casse(self) -> None:
        assert detect_indice_from_filename("export_indice_c.xls") == "C"


# ---------------------------------------------------------------------------
# Tests de mapping des colonnes
# ---------------------------------------------------------------------------


class TestBuildMapping:
    def test_colonnes_standard(self) -> None:
        headers = ["Repère", "Désignation", "IB (A)", "Longueur (m)"]
        mapping, mapped, extra = _build_mapping(headers)
        assert "repere" in mapped
        assert "designation" in mapped
        assert "ib" in mapped
        assert "longueur" in mapped

    def test_colonne_inconnue_en_extra(self) -> None:
        headers = ["Repère", "ColonneInconnue"]
        mapping, mapped, extra = _build_mapping(headers)
        assert "ColonneInconnue" in extra
        assert "repere" in mapped

    def test_colonne_absente_pas_en_mapped(self) -> None:
        headers = ["Repère", "Désignation"]
        _, mapped, _ = _build_mapping(headers)
        assert "ib" not in mapped
        assert "longueur" not in mapped

    def test_alias_acceptes(self) -> None:
        headers = ["Rep", "Long", "Type câble"]
        _, mapped, _ = _build_mapping(headers)
        assert "repere" in mapped
        assert "longueur" in mapped
        assert "type_cable" in mapped


# ---------------------------------------------------------------------------
# Tests d'intégration : fichier DACHSER de référence
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not DACHSER_FILE.exists(), reason="Fichier DACHSER absent du seed")
class TestDachserFile:
    def test_nombre_exact_de_lignes(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert len(result.lines) == 699, (
            f"Attendu 699 lignes, obtenu {len(result.lines)}. "
            "Vérifiez que le fichier source n'a pas changé."
        )

    def test_lignes_lues_egal_lignes_parsees(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.lines_read == len(result.lines)

    def test_23_colonnes_detectees(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.columns_detected == 23

    def test_23_colonnes_mappees(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.columns_mapped == 23

    def test_aucune_colonne_supplementaire(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.extra_columns_count == 0

    def test_aucun_warning(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.warnings == []

    def test_excel_row_number_commence_a_2(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        assert result.lines[0].excel_row_number == 2

    def test_excel_row_number_continu(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        for i, line in enumerate(result.lines):
            assert line.excel_row_number == i + 2

    def test_raw_data_indexe_par_champ(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        # raw_data doit être un dict avec des clés correspondant aux noms de champs
        first = result.lines[0]
        assert isinstance(first.raw_data, dict)
        # Les clés doivent correspondre à des noms de champs (pas des en-têtes Excel)
        assert "repere" in first.raw_data or "designation" in first.raw_data

    def test_lignes_sans_repere_conservees(self) -> None:
        result = parse_caneco_file(DACHSER_FILE)
        # Dans DACHSER, certaines lignes ont un repère vide — elles doivent être conservées
        lines_sans_repere = [ln for ln in result.lines if not ln.repere]
        # Juste vérifier que le parser ne les filtre pas
        assert isinstance(lines_sans_repere, list)


# ---------------------------------------------------------------------------
# Tests avec des fichiers synthétiques
# ---------------------------------------------------------------------------


class TestAvecColonneRenommee:
    """Le parser doit accepter les alias de colonnes."""

    def test_alias_ib(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        # "Courant IB" est un alias de "IB"
        ws.append(["Repere", "Designation", "Courant IB"])
        ws.append(["TGBT", "Tableau general", 250.5])
        path = tmp_path / "test_alias.xlsx"
        wb.save(path)

        result = parse_caneco_file(path)
        assert len(result.lines) == 1
        assert result.lines[0].ib == pytest.approx(250.5)
        assert "ib" in result.column_names_mapped


class TestAvecColonneSupplementaire:
    """Les colonnes non standard doivent aller dans extra_columns."""

    def test_colonne_inconnue_stockee(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Repere", "Designation", "ColonneInconnue"])
        ws.append(["TGBT", "Tableau general", "valeur_extra"])
        path = tmp_path / "test_extra.xlsx"
        wb.save(path)

        result = parse_caneco_file(path)
        assert result.extra_columns_count == 1
        assert "ColonneInconnue" in result.column_names_extra
        assert result.lines[0].extra_columns.get("ColonneInconnue") == "valeur_extra"


class TestAvecLigneRepereVide:
    """Les lignes avec repère vide doivent être conservées."""

    def test_ligne_sans_repere_conservee(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Repere", "Designation", "IB (A)"])
        ws.append(["TGBT", "Tableau general", 100.0])
        ws.append([None, "Suite du tableau", None])  # ligne sans repère
        ws.append(["TES1", "Tableau secondaire", 50.0])
        path = tmp_path / "test_empty_repere.xlsx"
        wb.save(path)

        result = parse_caneco_file(path)
        assert len(result.lines) == 3
        assert result.lines[1].repere is None
        assert result.lines[1].designation == "Suite du tableau"

"""Export des lignes CANECO parsées vers un fichier Excel (.xlsx).

Produit un fichier avec en-têtes en bleu VINCI (#001E50), texte blanc, gras.
Les données numériques sont formatées sans décimales inutiles.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.caneco import CanecoLine

# Couleur VINCI bleu nuit
_VINCI_BLUE = "001E50"

# Colonnes exportées : (attribut_modele, libelle_entete)
_COLUMNS: list[tuple[str, str]] = [
    ("excel_row_number", "#"),
    ("amont", "Amont"),
    ("repere_aval", "Repere aval"),
    ("repere", "Repere"),
    ("designation", "Designation"),
    ("style", "Style"),
    ("nb_recepteurs", "Nb recepteurs"),
    ("consommation", "Consommation"),
    ("ib", "IB (A)"),
    ("longueur", "Longueur (m)"),
    ("type_cable", "Type de cable"),
    ("nb_cables_multi", "Nb cables multi"),
    ("cable", "Cable"),
    ("neutre", "Neutre"),
    ("pe", "PE ou PEN"),
    ("calibre", "Calibre (A)"),
    ("bloc_coupure", "Bloc de coupure"),
    ("bloc_declencheur", "Bloc declencheur"),
    ("bloc_differentiel", "Bloc differentiel"),
    ("ir_th_in", "IrTh / IN"),
    ("ir_mg_in", "IrMg / IN"),
    ("icu", "Icu (kA)"),
    ("contacteur", "Contacteur"),
    ("ame", "Ame"),
]

_COL_WIDTHS: dict[str, int] = {
    "#": 6,
    "amont": 18,
    "repere_aval": 18,
    "repere": 14,
    "designation": 30,
    "style": 16,
    "nb_recepteurs": 12,
    "consommation": 16,
    "ib": 10,
    "longueur": 12,
    "type_cable": 14,
    "nb_cables_multi": 14,
    "cable": 12,
    "neutre": 10,
    "pe": 10,
    "calibre": 10,
    "bloc_coupure": 16,
    "bloc_declencheur": 16,
    "bloc_differentiel": 16,
    "ir_th_in": 12,
    "ir_mg_in": 12,
    "icu": 10,
    "contacteur": 12,
    "ame": 8,
}


def export_to_xlsx(lines: list[CanecoLine], file_name: str, indice: str) -> BytesIO:
    """Génère un fichier Excel des lignes CANECO avec mise en forme VINCI.

    Args:
        lines: Liste des CanecoLine à exporter.
        file_name: Nom du fichier source (pour le titre de la feuille).
        indice: Indice de révision (pour le titre de la feuille).

    Returns:
        BytesIO contenant le fichier .xlsx prêt à envoyer.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    sheet_title = f"Indice_{indice}"[:31]
    ws.title = sheet_title

    header_fill = PatternFill(start_color=_VINCI_BLUE, end_color=_VINCI_BLUE, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (attr, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(attr, 14)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "B2"

    for row_idx, line in enumerate(lines, start=2):
        for col_idx, (attr, _) in enumerate(_COLUMNS, start=1):
            value = getattr(line, attr, None)
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

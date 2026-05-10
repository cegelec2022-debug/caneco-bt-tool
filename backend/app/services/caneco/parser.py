"""Parser pour les exports CANECO BT au format XLS/XLSX.

Lit la première ligne comme en-tête, mappe les 23 colonnes standard par
normalisation des noms (accents, unités, casse), conserve TOUTES les lignes
(y compris celles avec un repère vide), stocke les valeurs brutes dans raw_data.
"""

import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from loguru import logger

# ---------------------------------------------------------------------------
# Correspondances colonnes CANECO → champs modèle
# ---------------------------------------------------------------------------

# Chaque entrée : (nom_normalise_possible, ...) → nom_champ_modele
# L'ordre des alternatives va du plus spécifique au plus générique.
_COLUMN_ALIASES: list[tuple[list[str], str]] = [
    (["amont"], "amont"),
    (["repere aval", "rep aval", "aval"], "repere_aval"),
    (["repere", "rep", "reference", "ref depart"], "repere"),
    (["designation", "libelle", "intitule", "desc"], "designation"),
    (["style", "type circuit", "type de circuit"], "style"),
    (["nb recepteurs", "nb rec", "nombre recepteurs", "nombre de recepteurs", "n rec"], "nb_recepteurs"),
    (["consommation", "conso", "puissance", "puissance installee"], "consommation"),
    (["ib", "courant calcule", "courant de calcul", "courant ib", "i calcule"], "ib"),
    (["longueur", "long", "lg"], "longueur"),
    (["type de cable", "type cable", "nature cable", "nature du cable"], "type_cable"),
    (["nb cables multi", "nb cable multi", "nb multiconducteurs", "nb multi"], "nb_cables_multi"),
    (["cable", "section cable", "section du cable", "section ph", "section phase", "section"], "cable"),
    (["neutre", "section neutre"], "neutre"),
    (["pe ou pen", "pe/pen", "pe", "section pe", "terre"], "pe"),
    (["calibre", "calibre disjoncteur", "calibre dj", "in dj", "in disjoncteur"], "calibre"),
    (["bloc de coupure", "bloc coupure", "coupure", "sectionneur"], "bloc_coupure"),
    (["bloc declencheur", "declencheur", "relais thermique"], "bloc_declencheur"),
    (["bloc differentiel", "differentiel", "diff", "ido"], "bloc_differentiel"),
    (["irth / in", "irth/in", "ir th / in", "ir th/in", "ir th in", "irth in", "reglage thermique"], "ir_th_in"),
    (["irmg / in", "irmg/in", "ir mg / in", "ir mg/in", "ir mg in", "irmg in", "reglage magnetique"], "ir_mg_in"),
    (["icu", "pouvoir de coupure", "pdc"], "icu"),
    (["contacteur", "contact"], "contacteur"),
    (["ame", "nb ames", "nombre ames", "nb conducteurs"], "ame"),
]

# Champs numériques (float)
_FLOAT_FIELDS = {"ib", "longueur", "calibre", "ir_th_in", "ir_mg_in", "icu"}

# Champs numériques (int)
_INT_FIELDS = {"nb_recepteurs", "nb_cables_multi"}

# Champs texte
_STR_FIELDS = {
    "amont", "repere_aval", "repere", "designation", "style", "consommation",
    "type_cable", "cable", "neutre", "pe", "bloc_coupure", "bloc_declencheur",
    "bloc_differentiel", "contacteur", "ame",
}

# Ensemble des champs standard pour détecter les colonnes supplémentaires
_STANDARD_FIELDS = _FLOAT_FIELDS | _INT_FIELDS | _STR_FIELDS


# ---------------------------------------------------------------------------
# Structures de données
# ---------------------------------------------------------------------------


@dataclass
class ParsedLine:
    """Représente une ligne parsée depuis un export CANECO BT."""

    excel_row_number: int
    row_index: int

    # Champs modèle
    amont: str | None = None
    repere_aval: str | None = None
    repere: str | None = None
    designation: str | None = None
    style: str | None = None
    nb_recepteurs: int | None = None
    consommation: str | None = None
    ib: float | None = None
    longueur: float | None = None
    type_cable: str | None = None
    nb_cables_multi: int | None = None
    cable: str | None = None
    neutre: str | None = None
    pe: str | None = None
    calibre: float | None = None
    bloc_coupure: str | None = None
    bloc_declencheur: str | None = None
    bloc_differentiel: str | None = None
    ir_th_in: float | None = None
    ir_mg_in: float | None = None
    icu: float | None = None
    contacteur: str | None = None
    ame: str | None = None

    # Valeurs brutes ({nom_colonne_excel: valeur_brute})
    raw_data: dict[str, str] = field(default_factory=dict)

    # Colonnes non reconnues ({nom_colonne_excel: valeur})
    extra_columns: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParseResult:
    """Résultat complet d'un parsing CANECO BT."""

    lines: list[ParsedLine]
    header_row_index: int
    lines_read: int
    columns_detected: int
    columns_mapped: int
    extra_columns_count: int
    warnings: list[str] = field(default_factory=list)
    column_names_detected: list[str] = field(default_factory=list)
    column_names_mapped: list[str] = field(default_factory=list)
    column_names_extra: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Normalisation des noms de colonnes
# ---------------------------------------------------------------------------


def _normalize(text: str) -> str:
    """Normalise un en-tête de colonne pour la correspondance.

    Exemple : "IrTh / IN" → "irth / in" → "irth / in"
    Exemple : "Repère aval" → "repere aval"
    """
    if not isinstance(text, str):
        text = str(text)
    # Suppression des accents (NFD → ASCII)
    nfd = unicodedata.normalize("NFD", text)
    ascii_text = nfd.encode("ascii", "ignore").decode("ascii")
    # Minuscules
    lower = ascii_text.lower()
    # Suppression des unités entre parenthèses : (A), (m), (kVA)
    no_units = re.sub(r"\([^)]*\)", "", lower)
    # Normalisation des espaces
    return " ".join(no_units.split())


def _build_mapping(header: list[str]) -> tuple[dict[int, str], list[str], list[str]]:
    """Construit le mapping {index_colonne: nom_champ_modele} depuis la ligne d'en-tête.

    Retourne (mapping, champs_mappes, colonnes_extra).
    """
    mapping: dict[int, str] = {}
    mapped_fields: set[str] = set()
    extra_col_names: list[str] = []

    normalized_headers = [(i, h, _normalize(h)) for i, h in enumerate(header) if h]

    for col_idx, original_name, norm in normalized_headers:
        matched_field: str | None = None

        # 1. Match exact : le nom normalisé est dans la liste des alias
        for aliases, field_name in _COLUMN_ALIASES:
            if field_name in mapped_fields:
                continue
            if norm in aliases:
                matched_field = field_name
                break

        # 2. Match partiel : un alias est contenu dans le nom normalisé
        #    (ex. header "ib courant calcule" contient l'alias "ib")
        #    On n'utilise PAS norm-in-alias car cela crée des faux positifs
        #    (ex. "repere" serait contenu dans "repere aval")
        if not matched_field:
            for aliases, field_name in _COLUMN_ALIASES:
                if field_name in mapped_fields:
                    continue
                for alias in aliases:
                    if alias in norm:
                        matched_field = field_name
                        break
                if matched_field:
                    break

        if matched_field:
            mapping[col_idx] = matched_field
            mapped_fields.add(matched_field)
        else:
            extra_col_names.append(original_name)

    return mapping, sorted(mapped_fields), extra_col_names


# ---------------------------------------------------------------------------
# Lecture des fichiers
# ---------------------------------------------------------------------------


def _read_xls(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    """Lit un .xls avec xlrd. Retourne (headers, rows_brutes)."""
    import xlrd  # type: ignore[import-untyped]

    wb = xlrd.open_workbook(str(file_path))
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        raise ValueError("Fichier XLS vide.")

    # Ligne 0 = en-têtes
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]

    # Lignes 1..N = données
    raw_rows: list[list[Any]] = []
    for r in range(1, ws.nrows):
        row: list[Any] = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            # ctype: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
            if cell.ctype == 0:
                row.append(None)
            elif cell.ctype == 2:
                # Nombre : garder float (même si entier, pour la précision)
                row.append(cell.value)
            else:
                text = str(cell.value).strip()
                row.append(text if text else None)
        raw_rows.append(row)

    return headers, raw_rows


def _read_xlsx(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    """Lit un .xlsx avec openpyxl. Retourne (headers, rows_brutes)."""
    import openpyxl  # type: ignore[import-untyped]

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
    wb.close()

    if not all_rows:
        raise ValueError("Fichier XLSX vide.")

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    raw_rows = [list(row) for row in all_rows[1:]]
    return headers, raw_rows


def _read_file(file_path: Path) -> tuple[list[str], list[list[Any]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(file_path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(file_path)
    raise ValueError(f"Format non supporté : {suffix}. Utilisez .xls ou .xlsx.")


# ---------------------------------------------------------------------------
# Conversion de types
# ---------------------------------------------------------------------------


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", ".").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    f = _to_float(value)
    if f is None:
        return None
    return int(round(f))


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in ("none", "nan", "#n/a", "n/a"):
        return None
    return text


def _raw_str(value: Any) -> str:
    """Valeur brute sous forme de string, toujours définie."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


# ---------------------------------------------------------------------------
# Parsing d'une ligne de données
# ---------------------------------------------------------------------------


def _parse_row(
    raw_row: list[Any],
    headers: list[str],
    col_mapping: dict[int, str],
    excel_row_number: int,
    row_index: int,
) -> ParsedLine:
    """Parse une ligne brute Excel en ParsedLine.

    Conserve toutes les lignes (même celles avec repère vide).
    """
    line = ParsedLine(excel_row_number=excel_row_number, row_index=row_index)
    raw_data: dict[str, str] = {}
    extra_columns: dict[str, Any] = {}

    for col_idx, value in enumerate(raw_row):
        if col_idx >= len(headers):
            break

        header_name = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
        raw_str = _raw_str(value)

        field_name = col_mapping.get(col_idx)
        if field_name is None:
            # Colonne supplémentaire non reconnue
            if value is not None:
                extra_columns[header_name] = _to_str(value) or raw_str
            continue

        # Valeur brute indexée par nom de champ (facilite la recherche côté client)
        raw_data[field_name] = raw_str

        if field_name in _FLOAT_FIELDS:
            setattr(line, field_name, _to_float(value))
        elif field_name in _INT_FIELDS:
            setattr(line, field_name, _to_int(value))
        else:  # _STR_FIELDS
            setattr(line, field_name, _to_str(value))

    line.raw_data = raw_data
    line.extra_columns = extra_columns
    return line


# ---------------------------------------------------------------------------
# Détection automatique de l'indice depuis le nom de fichier
# ---------------------------------------------------------------------------


def detect_indice_from_filename(filename: str) -> str | None:
    """Extrait l'indice de révision depuis le nom du fichier.

    Exemples :
      "DATA_DACHSER_INDICE_B.XLS" → "B"
      "EXPORT CANECO INDICE B2.xlsx" → "B2"
      "projet_v3.xls" → "3" (fallback sur numéro de version)
    """
    # Pattern : INDICE suivi de lettre(s) + chiffre(s) optionnels
    m = re.search(r"INDICE[_\s]+([A-Z][0-9]*)", filename.upper())
    if m:
        return m.group(1)
    # Fallback : lettre unique après underscore ou espace (ex. "_B.", " B.")
    m = re.search(r"[_\s]([A-Z])\.", filename.upper())
    if m:
        return m.group(1)
    return None


# ---------------------------------------------------------------------------
# Fonction principale
# ---------------------------------------------------------------------------


def parse_caneco_file(file_path: str | Path) -> ParseResult:
    """Parse un export CANECO BT (XLS ou XLSX).

    Args:
        file_path: Chemin vers le fichier Excel CANECO BT.

    Returns:
        ParseResult avec toutes les lignes, y compris celles avec repère vide.

    Raises:
        ValueError: Si le fichier est invalide ou non reconnu comme export CANECO.
        FileNotFoundError: Si le fichier n'existe pas.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    logger.info(f"Début du parsing CANECO : {path.name}")

    headers, raw_rows = _read_file(path)

    # Vérification minimale : au moins 3 colonnes et au moins une ligne de données
    if len(headers) < 3:
        raise ValueError(
            f"Le fichier ne contient que {len(headers)} colonnes. "
            "Vérifiez qu'il s'agit bien d'un export CANECO BT."
        )
    if not raw_rows:
        raise ValueError("Le fichier ne contient aucune ligne de données sous l'en-tête.")

    col_mapping, mapped_fields, extra_col_names = _build_mapping(headers)

    if not col_mapping:
        raise ValueError(
            "Aucune colonne CANECO reconnue dans l'en-tête. "
            f"En-têtes trouvées : {headers[:5]}..."
        )

    logger.debug(
        f"En-tête : {len(headers)} colonnes — "
        f"{len(col_mapping)} mappées, {len(extra_col_names)} supplémentaires"
    )

    # Avertissements pour les colonnes standard non trouvées
    warnings: list[str] = []
    all_standard_fields = {f for _, f in _COLUMN_ALIASES}
    missing = all_standard_fields - set(mapped_fields)
    for f in sorted(missing):
        msg = f"Colonne standard '{f}' non trouvée dans cet export — champ NULL en base."
        logger.warning(f"[CANECO parser] {msg}")
        warnings.append(msg)

    # Parsing de TOUTES les lignes (sans filtrage)
    parsed_lines: list[ParsedLine] = []
    for row_idx, raw_row in enumerate(raw_rows):
        # excel_row_number = 2 pour la première donnée (ligne 1 = en-tête)
        excel_row = row_idx + 2
        parsed = _parse_row(raw_row, headers, col_mapping, excel_row, row_idx)
        parsed_lines.append(parsed)

    logger.info(
        f"Parsing terminé : {len(parsed_lines)}/{len(raw_rows)} lignes traitées "
        f"({len(col_mapping)}/{len(headers)} colonnes mappées) — {path.name}"
    )

    return ParseResult(
        lines=parsed_lines,
        header_row_index=0,
        lines_read=len(raw_rows),
        columns_detected=len(headers),
        columns_mapped=len(col_mapping),
        extra_columns_count=len(extra_col_names),
        warnings=warnings,
        column_names_detected=headers,
        column_names_mapped=mapped_fields,
        column_names_extra=extra_col_names,
    )

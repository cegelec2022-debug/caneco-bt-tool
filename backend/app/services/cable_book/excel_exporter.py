"""Generation du fichier Excel du carnet de cables (feuilles Sommaire + Rapport).

Utilise openpyxl (deja dans le projet pour le parsing CANECO).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.cable_book.builder import CableBookReport


# Styles VINCI
_HEADER_FILL = PatternFill(start_color="001E50", end_color="001E50", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="001E50")
_SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="404040")
_BODY_FONT = Font(name="Calibri", size=10)
_MONO_FONT = Font(name="Consolas", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def build_cable_book_workbook(
    report: CableBookReport,
    *,
    project_name: str,
    project_code: str,
    indice: str = "",
) -> bytes:
    """Genere le fichier Excel du carnet de cables.

    Args:
        report: CableBookReport produit par build_cable_book().
        project_name: Nom complet du projet (pour le titre).
        project_code: Code du projet.
        indice: Indice de l'export CANECO source (optionnel).

    Returns:
        Bytes du fichier .xlsx.
    """
    wb = Workbook()
    # Premiere feuille : Sommaire
    ws_summary = wb.active
    ws_summary.title = "Sommaire Cables"
    _build_summary_sheet(ws_summary, report, project_name, project_code, indice)

    # Deuxieme feuille : Rapport
    ws_report = wb.create_sheet("Rapport Cables")
    _build_report_sheet(ws_report, report, project_name, project_code, indice)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(
    ws,
    report: CableBookReport,
    project_name: str,
    project_code: str,
    indice: str,
) -> None:
    """Construit la feuille Sommaire (1 ligne par type+section, triee par longueur)."""
    # Titre
    ws["A1"] = f"Carnet de cables — {project_name}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = f"Projet : {project_code}" + (f" — Indice {indice}" if indice else "")
    ws["A2"].font = _SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    # Stats rapides en haut
    ws["A4"] = "Longueur totale projet :"
    ws["A4"].font = _SUBTITLE_FONT
    ws["B4"] = round(report.longueur_totale_projet_m, 1)
    ws["B4"].number_format = "#,##0.0 \"m\""

    ws["A5"] = "Nb lignes CANECO traitees :"
    ws["A5"].font = _SUBTITLE_FONT
    ws["B5"] = report.nb_lignes_caneco_traitees

    ws["A6"] = "Nb types de cables distincts :"
    ws["A6"].font = _SUBTITLE_FONT
    ws["B6"] = report.nb_types_cables_distincts

    # En-tete du tableau
    header_row = 8
    headers = [
        "Type de cable",
        "Section CANECO",
        "Section (mm²)",
        "Longueur totale (m)",
        "Nb conducteurs",
        "Nb occurrences",
        "% du total projet",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    # Lignes
    for row_idx, entry in enumerate(report.entries, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=entry.type_cable).font = _BODY_FONT
        cell_cable = ws.cell(row=row_idx, column=2, value=entry.cable_caneco)
        cell_cable.font = _MONO_FONT
        ws.cell(row=row_idx, column=3, value=entry.section_mm2)
        ws.cell(row=row_idx, column=4, value=round(entry.longueur_totale_m, 1))
        ws.cell(row=row_idx, column=5, value=entry.nb_conducteurs)
        ws.cell(row=row_idx, column=6, value=entry.nb_occurrences)
        ws.cell(row=row_idx, column=7, value=round(entry.pourcentage_du_total, 2))

        for col_idx in range(1, 8):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = _THIN_BORDER
            if col_idx >= 3:
                c.alignment = _RIGHT
            else:
                c.alignment = _LEFT
            if col_idx not in (1, 2):
                c.font = _BODY_FONT

        ws.cell(row=row_idx, column=4).number_format = "#,##0.0"
        ws.cell(row=row_idx, column=7).number_format = "0.00\"%\""

    # Largeurs des colonnes
    widths = [22, 22, 14, 20, 16, 16, 18]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Fige les volets
    ws.freeze_panes = f"A{header_row + 1}"


def _build_report_sheet(
    ws,
    report: CableBookReport,
    project_name: str,
    project_code: str,
    indice: str,
) -> None:
    """Construit la feuille Rapport (KPIs, Top 5, longueur par tableau aval)."""
    ws["A1"] = f"Rapport synthetique cables — {project_name}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Projet : {project_code}" + (f" — Indice {indice}" if indice else "")
    ws["A2"].font = _SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    # KPIs
    row = 4
    ws.cell(row=row, column=1, value="INDICATEURS GENERAUX").font = _TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Longueur totale projet (m)").font = _SUBTITLE_FONT
    c = ws.cell(row=row, column=2, value=round(report.longueur_totale_projet_m, 1))
    c.number_format = "#,##0.0"
    row += 1
    ws.cell(row=row, column=1, value="Nb lignes CANECO traitees").font = _SUBTITLE_FONT
    ws.cell(row=row, column=2, value=report.nb_lignes_caneco_traitees)
    row += 1
    ws.cell(row=row, column=1, value="Nb types de cables distincts").font = _SUBTITLE_FONT
    ws.cell(row=row, column=2, value=report.nb_types_cables_distincts)
    row += 2

    # Top 5
    ws.cell(row=row, column=1, value="TOP 5 DES CABLES LES PLUS UTILISES").font = _TITLE_FONT
    row += 1
    headers = ["Rang", "Type cable", "Section CANECO", "Longueur (m)", "% projet"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    row += 1
    for rank, entry in enumerate(report.top5, start=1):
        ws.cell(row=row, column=1, value=rank).alignment = _CENTER
        ws.cell(row=row, column=2, value=entry.type_cable)
        cell_cable = ws.cell(row=row, column=3, value=entry.cable_caneco)
        cell_cable.font = _MONO_FONT
        c_len = ws.cell(row=row, column=4, value=round(entry.longueur_totale_m, 1))
        c_len.number_format = "#,##0.0"
        c_pct = ws.cell(row=row, column=5, value=round(entry.pourcentage_du_total, 2))
        c_pct.number_format = "0.00\"%\""
        for col_idx in range(1, 6):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1

    row += 1

    # Longueur par type
    ws.cell(row=row, column=1, value="LONGUEUR PAR TYPE DE CABLE").font = _TITLE_FONT
    row += 1
    headers = ["Type cable", "Longueur (m)", "% projet"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    row += 1
    total = report.longueur_totale_projet_m or 1
    for tc, lg in sorted(
        report.longueur_par_type_cable.items(), key=lambda x: x[1], reverse=True
    ):
        ws.cell(row=row, column=1, value=tc)
        c_len = ws.cell(row=row, column=2, value=round(lg, 1))
        c_len.number_format = "#,##0.0"
        c_pct = ws.cell(row=row, column=3, value=round(lg / total * 100, 2))
        c_pct.number_format = "0.00\"%\""
        for col_idx in range(1, 4):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1

    row += 1

    # Longueur par tableau aval
    if report.longueur_par_aval:
        ws.cell(row=row, column=1, value="LONGUEUR PAR TABLEAU AVAL / LOT").font = _TITLE_FONT
        row += 1
        headers = ["Tableau aval", "Longueur (m)", "% projet"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER
        row += 1
        for aval, lg in sorted(
            report.longueur_par_aval.items(), key=lambda x: x[1], reverse=True
        ):
            ws.cell(row=row, column=1, value=aval)
            c_len = ws.cell(row=row, column=2, value=round(lg, 1))
            c_len.number_format = "#,##0.0"
            c_pct = ws.cell(row=row, column=3, value=round(lg / total * 100, 2))
            c_pct.number_format = "0.00\"%\""
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = _THIN_BORDER
            row += 1

    # Largeurs
    widths = [30, 22, 18, 18, 14]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

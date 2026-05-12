"""Parser du bordereau de prix — feuille choisie par l'utilisateur ou auto-detectee.

Logique de detection de feuille :
  1. Si sheet_name est fourni → utiliser directement cette feuille
  2. Sinon → chercher une feuille dont le nom contient "ELECTRICITE" et "CFO"
  3. Sinon → chercher une feuille contenant "ELECTRICITE"
  4. Sinon → utiliser la premiere feuille qui n'est pas de type recap/fluide/IT
  5. Sinon → lever une erreur listant les feuilles disponibles

Logique de parsing adaptative :
  - En-tete : ligne contenant "N°PRIX" (ou variantes)
  - Sections principales : col A matche \d{3,4}-.+ (ex. "500-ELECTRICITE CFO")
  - Sous-sections    : col A matche \d{2,4} seul (ex. "505", "50")
  - Articles         : col A matche \d+\.\d+ (ex. "505.3", "1.1", "5001.2")
  - Sous-familles    : col A vide, col B contient texte (titre intermediaire)
  - Fallback section : si aucune section detectee, creer une section "000 - General"

Enrichissement des cables (section 505 ou type detecte "cable") :
  - detected_section_mm2 : "5G6", "1X240", "4X95+T50"
  - detected_material    : "ALU" ou "CUIVRE"
  - detected_cable_type  : "U1000AR2V", "U1000R2V", "CR1", "FR-N1X1", "H07VR"
  - detected_kind        : "cable", "tableau", "chemin_cable", "tubage", etc.
"""

import re
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from app.models.bordereau import BordereauImport, BordereauLine, BordereauSection

# ---------------------------------------------------------------------------
# Patterns de reconnaissance des lignes
# ---------------------------------------------------------------------------

# Section principale : "500-ELECTRICITE COURANT FORT", "5000-ELECTRICITE"
_RE_SECTION_MAIN = re.compile(r"^\d{3,4}-.+")

# Sous-section : "505", "50", "5005" (2 a 4 chiffres seuls)
_RE_SECTION_SUB = re.compile(r"^\d{2,4}$")

# Article : tout pattern "code.numero" (ex. "505.3", "1.1", "5001.2", "A.1")
_RE_ARTICLE = re.compile(r"^[\w]+\.\d+$")

# ---------------------------------------------------------------------------
# Patterns d'enrichissement cables
# ---------------------------------------------------------------------------

# Section conducteur : "5G6", "1X240", "4X95+T50", "3x1.5", "2x2.5"
_RE_SECTION_MM2 = re.compile(
    r"(\d+\s*[xXgG×]\s*\d+(?:[.,]\d+)?(?:\s*[+]\s*T\s*\d+)?)",
    re.IGNORECASE,
)

# Type de cable : U1000R2V, U1000AR2V, U1000RO2V, U1000ARO2V, CR1, FR-N1X1, H07VR
_RE_CABLE_TYPE = re.compile(
    r"(U1000\s*A?\s*R\s*O?\s*2V|CR1|FR-N1X1|H07[VRU][RN]?-?[FKR]?)",
    re.IGNORECASE,
)

# Feuilles a ignorer lors de l'auto-detection (fluides, IT, recap)
_SHEET_IGNORE_KEYWORDS = {"FLUIDE", "PLOMBERIE", "RECAP", "IT", "PRECABLAGE", "CFA", "SURETE"}

# Mapping code section → detected_kind
_SECTION_KIND_MAP: dict[str, str] = {
    "505": "cable",
    "506": "chemin_cable",
    "507": "tableau",
    "508": "tubage",
    "509": "appareillage",
    "519": "paratonnerre",
    "520": "parafoudre",
}


# ---------------------------------------------------------------------------
# Utilitaires
# ---------------------------------------------------------------------------


def _cell_str(cell: Any) -> str | None:
    """Retourne la valeur d'une cellule comme chaine, ou None si vide."""
    if cell is None or cell.value is None:
        return None
    val = str(cell.value).strip()
    return val if val else None


def _is_all_caps_title(text: str) -> bool:
    """Retourne True si le texte ressemble a un titre de section en majuscules."""
    stripped = re.sub(r"[^A-Za-z]", "", text)
    if len(stripped) < 4:
        return False
    return stripped == stripped.upper()


# ---------------------------------------------------------------------------
# Detection de la feuille
# ---------------------------------------------------------------------------


def list_sheet_names(file_content: bytes) -> tuple[list[str], str | None]:
    """Lit les noms de feuilles d'un fichier Excel et retourne la feuille recommandee.

    Args:
        file_content: Contenu binaire du fichier Excel.

    Returns:
        Tuple (liste_de_feuilles, feuille_recommandee_ou_None).
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()

    detected = _detect_cfo_sheet_name(sheets)
    return sheets, detected


def _detect_cfo_sheet_name(sheet_names: list[str]) -> str | None:
    """Trouve le nom de la feuille BDP_ELECTRICITE CFO parmi une liste."""
    # Passage 1 : electricite + CFO
    for name in sheet_names:
        normalized = re.sub(r"\s+", "", name).upper()
        if "ELECTRICITE" in normalized and "CFO" in normalized:
            return name
    # Passage 2 : electricite seul
    for name in sheet_names:
        normalized = re.sub(r"\s+", "", name).upper()
        if "ELECTRICITE" in normalized:
            return name
    # Passage 3 : premiere feuille qui n'est pas a ignorer
    for name in sheet_names:
        upper = re.sub(r"\s+", "", name).upper()
        if not any(kw in upper for kw in _SHEET_IGNORE_KEYWORDS):
            return name
    return None


def _find_sheet(wb: openpyxl.Workbook, sheet_name: str | None) -> Any:
    """Retourne la feuille demandee ou auto-detectee.

    Args:
        wb: Workbook deja ouvert.
        sheet_name: Nom de feuille explicite (None = auto-detection).

    Raises:
        ValueError: Si la feuille n'est pas trouvee.
    """
    if sheet_name:
        if sheet_name in wb.sheetnames:
            logger.info(f"Feuille selectionnee : '{sheet_name}'")
            return wb[sheet_name]
        raise ValueError(
            f"Feuille '{sheet_name}' introuvable. "
            f"Feuilles disponibles : {wb.sheetnames}"
        )

    detected = _detect_cfo_sheet_name(list(wb.sheetnames))
    if detected:
        logger.info(f"Feuille auto-detectee : '{detected}'")
        return wb[detected]

    raise ValueError(
        f"Aucune feuille electrique detectee automatiquement. "
        f"Feuilles disponibles : {wb.sheetnames}. "
        f"Specifiez le nom de la feuille a analyser."
    )


# ---------------------------------------------------------------------------
# Detection de la ligne d'en-tete
# ---------------------------------------------------------------------------


def _find_header_row(ws: Any) -> int:
    """Trouve le numero de la ligne d'en-tete (contenant N°PRIX ou N° PRIX).

    Returns:
        Numero de ligne (1-indexed).

    Raises:
        ValueError: Si l'en-tete n'est pas trouve dans les 20 premieres lignes.
    """
    for row_num in range(1, 21):
        for cell in ws[row_num]:
            val = _cell_str(cell)
            if val and re.search(r"N[°uo]?\s*PRIX", val, re.IGNORECASE):
                logger.info(f"Ligne d'en-tete trouvee a la ligne {row_num}")
                return row_num
    raise ValueError("En-tete 'N°PRIX' introuvable dans les 20 premieres lignes du fichier.")


# ---------------------------------------------------------------------------
# Enrichissement des lignes
# ---------------------------------------------------------------------------


def _detect_material(text: str) -> str | None:
    """Detecte le materiau conducteur depuis un texte."""
    upper = text.upper()
    if re.search(r"\bALU\b|ALUM", upper):
        return "ALU"
    if re.search(r"\bCUIVRE\b|CUIV\b", upper):
        return "CUIVRE"
    return None


def _detect_section_mm2(text: str) -> str | None:
    """Extrait la section conducteur depuis la designation."""
    m = _RE_SECTION_MM2.search(text)
    if m:
        return re.sub(r"\s+", "", m.group(1)).upper()
    return None


def _detect_cable_type(text: str) -> str | None:
    """Extrait le type de cable depuis la designation."""
    m = _RE_CABLE_TYPE.search(text)
    if m:
        return re.sub(r"\s+", "", m.group(1)).upper()
    return None


def _section_code_from_num_prix(num_prix: str) -> str | None:
    """Extrait le code section depuis un num_prix (ex. '505.3' → '505')."""
    parts = num_prix.split(".")
    return parts[0].strip() if parts else None


def _kind_from_section_code(code: str) -> str:
    """Retourne le detected_kind correspondant au code section."""
    return _SECTION_KIND_MAP.get(code, "autre")


def _enrich_line(line: BordereauLine, section_code: str | None, sous_famille: str | None) -> None:
    """Enrichit les champs detected_* d'une BordereauLine par detection regex."""
    designation = line.designation or ""
    context = " ".join(filter(None, [designation, sous_famille or ""]))

    line.detected_kind = _kind_from_section_code(section_code) if section_code else "autre"

    if line.detected_kind == "cable":
        line.detected_section_mm2 = _detect_section_mm2(context)
        line.detected_cable_type = _detect_cable_type(context)
        line.detected_material = _detect_material(context)
    else:
        line.detected_section_mm2 = None
        line.detected_cable_type = None
        line.detected_material = None


# ---------------------------------------------------------------------------
# Resultat du parsing
# ---------------------------------------------------------------------------


class ParseResult:
    """Resultat du parsing d'un fichier bordereau."""

    def __init__(self) -> None:
        self.sections: list[BordereauSection] = []
        self.lines: list[BordereauLine] = []
        self.total_lines: int = 0
        self.total_articles: int = 0
        self.sections_count: int = 0
        self.sheet_name_used: str = ""


# ---------------------------------------------------------------------------
# Point d'entree public
# ---------------------------------------------------------------------------


def parse_bordereau_file(
    file_path: Path,
    import_id: str,
    sheet_name: str | None = None,
) -> ParseResult:
    """Parse le fichier bordereau et retourne les sections et lignes extraites.

    Args:
        file_path: Chemin absolu vers le fichier xlsx.
        import_id: ID de l'import BordereauImport cible.
        sheet_name: Nom de la feuille a parser (None = auto-detection).

    Returns:
        ParseResult contenant les objets SQLAlchemy a persister.

    Raises:
        ValueError: Si la feuille est absente ou le format invalide.
    """
    logger.info(f"Parsing bordereau : {file_path} (feuille : {sheet_name or 'auto'})")

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    ws = _find_sheet(wb, sheet_name)
    header_row = _find_header_row(ws)

    result = ParseResult()
    result.sheet_name_used = ws.title  # type: ignore[assignment]

    current_section: BordereauSection | None = None
    current_sous_famille: str | None = None
    section_order = 0

    rows = list(ws.iter_rows(min_row=header_row + 1))
    result.total_lines = len(rows)

    for row in rows:
        excel_row_num = row[0].row if row else None

        col_a = _cell_str(row[0]) if len(row) > 0 else None
        col_b = _cell_str(row[1]) if len(row) > 1 else None
        col_c = _cell_str(row[2]) if len(row) > 2 else None
        col_d = _cell_str(row[3]) if len(row) > 3 else None

        # Ligne entierement vide → ignorer
        if not col_a and not col_b:
            continue

        # --- Detection du type de ligne ---

        # Section principale : "500-ELECTRICITE COURANT FORT"
        if col_a and _RE_SECTION_MAIN.match(col_a):
            dash_pos = col_a.index("-")
            code = col_a[:dash_pos].strip()
            title = col_a[dash_pos + 1:].strip()
            section = _make_section(import_id, code, title, excel_row_num, section_order)
            result.sections.append(section)
            current_section = section
            current_sous_famille = None
            section_order += 1
            continue

        # Sous-section : col A = "505" (2-4 chiffres seuls)
        if col_a and _RE_SECTION_SUB.match(col_a):
            title = col_b.strip() if col_b else None
            section = _make_section(import_id, col_a.strip(), title, excel_row_num, section_order)
            result.sections.append(section)
            current_section = section
            current_sous_famille = None
            section_order += 1
            continue

        # Article : col A = "505.3" ou "1.1" ou "5001.2"
        if col_a and _RE_ARTICLE.match(col_a):
            num_prix = col_a.strip()
            designation = col_b.strip() if col_b else None
            unite = col_c.strip().upper() if col_c else None
            quantite_raw = str(col_d).strip() if col_d else None
            quantite: float | None = None
            if quantite_raw:
                try:
                    quantite = float(str(quantite_raw).replace(",", "."))
                except (ValueError, TypeError):
                    quantite = None

            section_code = _section_code_from_num_prix(num_prix)

            line = BordereauLine(
                id=str(uuid.uuid4()),
                bordereau_import_id=import_id,
                section_id=current_section.id if current_section else None,
                excel_row_number=excel_row_num,
                num_prix=num_prix,
                designation=designation,
                unite=unite,
                quantite=quantite,
                quantite_raw=quantite_raw,
                sous_famille=current_sous_famille,
            )
            _enrich_line(line, section_code, current_sous_famille)

            result.lines.append(line)
            result.total_articles += 1
            continue

        # Sous-famille ou section implicite : col A vide, col B contient un titre
        if not col_a and col_b:
            text = col_b.strip()
            if _is_all_caps_title(text):
                # Titre en majuscules sans code → creer une section implicite
                auto_code = str(section_order).zfill(3)
                section = _make_section(import_id, auto_code, text, excel_row_num, section_order)
                result.sections.append(section)
                current_section = section
                current_sous_famille = None
                section_order += 1
            else:
                # Titre mixte → sous-famille
                current_sous_famille = text
            continue

    wb.close()

    # Fallback : si aucune section n'a ete detectee, creer une section "General"
    if not result.sections and result.lines:
        fallback_section = _make_section(import_id, "000", "General", None, 0)
        result.sections.append(fallback_section)
        for line in result.lines:
            line.section_id = fallback_section.id

    result.sections_count = len(result.sections)
    logger.info(
        f"Parsing termine : {result.total_lines} lignes lues, "
        f"{result.total_articles} articles, {result.sections_count} sections "
        f"(feuille : {result.sheet_name_used})"
    )
    return result


def _make_section(
    import_id: str,
    code: str,
    title: str | None,
    excel_row: int | None,
    order: int,
) -> BordereauSection:
    return BordereauSection(
        id=str(uuid.uuid4()),
        bordereau_import_id=import_id,
        code=code,
        title=title,
        excel_row_number=excel_row,
        order_index=order,
    )


def detect_indice_from_filename(filename: str) -> str | None:
    """Detecte l'indice de revision depuis le nom de fichier."""
    upper = filename.upper()
    m = re.search(r"INDICE[_\s]+([A-Z][0-9]*)", upper)
    if m:
        return m.group(1)
    m2 = re.search(r"[_\s]([A-Z])\.", upper)
    if m2:
        return m2.group(1)
    return None
